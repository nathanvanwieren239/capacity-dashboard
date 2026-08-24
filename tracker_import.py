"""
Import the real Gate Zero / Project Launch Tracker workbook.

    python tracker_import.py "path/to/tracker.xlsm"

Reads the **Project Launch Tracker** sheet and, where available, joins the
**Gate Zero Summary-NA&SA** sheet on customer part number to pick up the
launch process and a few sales-side fields.

--------------------------------------------------------------------------
COLUMN MAP - Project Launch Tracker
--------------------------------------------------------------------------
Row 2 holds group headers, row 3 holds sub-headers, data starts at row 4.

    A  Plant / Location          B  Div (MS/PS)        C  Customer name
    D  Customer part number      E  Description        F  Gate Zero RPN
    G  Gate Zero date            H  Project status R/Y/G
    I  Status (In-Process / Complete)
    J  Job number                K  Project manager    L  Peak annual sales
    M/N/O  PPAP  original / adjusted / actual
    P/Q/R  SOP   original / adjusted / actual
    S  QMSI capex   T  CER amount   U  CER status   V  CER number
    W  Project Initiation date
    X/Y/Z/AA     Simple Launch  plan / adjusted / actual / status
    AB/AC/AD/AE  Gate 1
    AF/AG/AH/AI  Gate 2
    AJ/AK/AL/AM  Gate 3   (mirrors PPAP)
    AN/AO/AP/AQ  Gate 4   (mirrors SOP)
    AR/AS/AT/AU  6 Month Review
    AV  PRR total amount 1st yr   AW  PRR start   AX  PRR end   AY  Notes

Verified against the workbook: Gate Zero date == Project Initiation date,
Gate 3 columns == PPAP columns, Gate 4 columns == SOP columns, and the
Simple Launch date == the PPAP date. Gate 1 and Gate 2 plan dates sit at
exactly one third and two thirds of kickoff-to-PPAP, which is why the
dashboard can recalculate them.

Nothing is written until the whole file parses. Rows without a part number
or description are skipped.
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import pandas as pd

import safe_io

DATA_DIR = Path(__file__).parent / "data"

TRACKER_SHEET = "Project Launch Tracker"
GATE_ZERO_SHEET = "Gate Zero Summary-NA&SA"
FIRST_DATA_ROW = 4

# gate_no, gate_code, gate_name, plan col, adjusted col, actual col, status col
GATE_MAP = [
    (0, "0", "Project Initiation", "W", None, None, None),
    (2, "SL", "Simple Launch", "X", "Y", "Z", "AA"),
    (1, "1", "Gate 1", "AB", "AC", "AD", "AE"),
    (2, "2", "Gate 2", "AF", "AG", "AH", "AI"),
    (3, "3", "Gate 3 — PPAP", "AJ", "AK", "AL", "AM"),
    (4, "4", "Gate 4 — SOP", "AN", "AO", "AP", "AQ"),
    (5, "6M", "6 Month Review", "AR", "AS", "AT", "AU"),
]

QA_HOURS = {"0": 2.0, "1": 6.0, "2": 10.0, "SL": 26.0, "3": 44.0, "4": 8.0, "6M": 4.0}

PROJECT_COLS = {
    "plant": "A", "div": "B", "customer": "C", "customer_part_number": "D",
    "description": "E", "rpn": "F", "gate_zero_date": "G",
    "project_status": "H", "project_phase": "I", "job_number": "J",
    "program_manager": "K", "peak_annual_sales": "L",
    "ppap_target_date": "M", "sop_target_date": "P",
    "qmsi_capex": "S", "cer_amount": "T", "cer_status": "U", "cer_number": "V",
    "prr_amount_first_year": "AV", "prr_start_date": "AW",
    "prr_end_date": "AX", "notes": "AY",
}

RYG = {"G": "Green", "Y": "Yellow", "R": "Red"}


class ImportError_(ValueError):
    pass


# Excel error values arrive as strings like '#REF!' or '#N/A'.
EXCEL_ERRORS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def _blank(v) -> bool:
    if v is None:
        return True
    s = str(v).strip().upper()
    if s.startswith("#") or s in EXCEL_ERRORS:
        return True
    return s in ("", "N/A", "NA", "NONE", "-", "TBD")


def _is_excel_error(v) -> bool:
    return v is not None and str(v).strip().upper().startswith("#")


def _date(v):
    if _blank(v):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    ts = pd.to_datetime(v, errors="coerce")
    return None if pd.isna(ts) else ts.date()


def _num(v, default=0.0) -> float:
    if _blank(v):
        return default
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return default


def _text(v) -> str:
    """Text form, with whole floats normalised so 6428265.0 joins to 6428265."""
    if _blank(v):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _gate_zero_lookup(wb) -> dict[str, dict]:
    """Part number -> sales-side fields from the Gate Zero Summary sheet."""
    if GATE_ZERO_SHEET not in wb.sheetnames:
        return {}
    ws = wb[GATE_ZERO_SHEET]
    out: dict[str, dict] = {}
    for row in range(3, ws.max_row + 1):
        part = _text(ws[f"G{row}"].value)
        if not part:
            continue
        out[part] = {
            "sales_person": _text(ws[f"D{row}"].value),
            "gate_zero_corp": _text(ws[f"E{row}"].value),
            "opportunity_number": _text(ws[f"J{row}"].value),
            "launch_process": _text(ws[f"N{row}"].value),
            "support_required": _text(ws[f"O{row}"].value),
            "main_risk_comments": _text(ws[f"S{row}"].value),
        }
    return out


def parse(path_or_buffer) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Return (projects, gates, warnings). Raises ImportError_ on bad input."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ImportError_(
            "openpyxl is required to read the tracker workbook. "
            "Run: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
    if TRACKER_SHEET not in wb.sheetnames:
        raise ImportError_(
            f"Workbook has no '{TRACKER_SHEET}' sheet. Found: "
            + ", ".join(wb.sheetnames)
        )

    ws = wb[TRACKER_SHEET]
    lookup = _gate_zero_lookup(wb)

    projects, gates, warnings = [], [], []
    seen: set[str] = set()

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        def cell(col):
            return ws[f"{col}{row}"].value

        part = _text(cell(PROJECT_COLS["customer_part_number"]))
        desc = _text(cell(PROJECT_COLS["description"]))
        job = _text(cell(PROJECT_COLS["job_number"]))
        pm = _text(cell(PROJECT_COLS["program_manager"]))

        # Columns A-G are formula driven. When those formulas are broken the
        # row still holds real work - job number, PM, gate dates, notes - so
        # it is imported and flagged rather than dropped.
        broken_identity = _is_excel_error(cell(PROJECT_COLS["customer_part_number"]))

        # Any date anywhere on the row, used as a last-resort kickoff.
        row_dates = [
            d for d in (
                _date(cell(c))
                for _, _, _, pc, ac, actc, _ in GATE_MAP
                for c in (pc, ac, actc) if c
            ) if d
        ]

        if not (part or desc or job or pm or row_dates):
            continue

        raw_gz = cell(PROJECT_COLS["gate_zero_date"])
        gate_zero = _date(raw_gz) or _date(cell("W"))
        recovered_kickoff = False
        if gate_zero is None and row_dates:
            gate_zero = min(row_dates)
            recovered_kickoff = True
        if gate_zero is None:
            warnings.append(
                f"Row {row}: no usable dates at all — skipped."
            )
            continue

        label = part or desc or job or f"row {row}"
        if broken_identity:
            warnings.append(
                f"Row {row} ({label}): identity columns are #REF! in this "
                "workbook — imported with whatever survived."
                + (" Kickoff taken from the earliest gate date." if recovered_kickoff else "")
            )

        key = part or desc or job or f"row{row}"
        pid = f"L-{key}"[:40]
        if pid in seen:
            pid = f"{pid}-{row}"
        seen.add(pid)

        extra = lookup.get(part, {})
        process = extra.get("launch_process", "")

        # Launch type: prefer the Gate Zero sheet, else infer from whether
        # gates 1 and 2 carry plan dates.
        if process.lower().startswith("simple"):
            launch_type = "Simple"
        elif process.lower().startswith("full"):
            launch_type = "Full"
        else:
            # Infer from Gate 1. Explicit "N/A" across its columns means the
            # project skips gates 1-3, so it is a simple launch. A real date
            # or status means it is full. #REF! tells us nothing either way.
            g1 = [cell(c) for c in ("AB", "AC", "AD", "AE")]
            g1_values = [str(c).strip().upper() for c in g1 if c is not None]
            g1_all_na = bool(g1_values) and all(v in ("N/A", "NA") for v in g1_values)
            g1_has_data = any(_date(c) for c in g1[:3]) or _text(g1[3]) != ""

            if g1_all_na:
                launch_type = "Simple"
            elif g1_has_data:
                launch_type = "Full"
            else:
                launch_type = "Full"
                warnings.append(
                    f"Row {row} ({part or desc or job}): launch process unknown "
                    "and Gate 1 columns are unreadable — assumed Full."
                )

        ryg = _text(cell(PROJECT_COLS["project_status"])).upper()
        projects.append(
            {
                "project_id": pid,
                "project_name": (
                    f"{part} — {desc}".strip(" —")
                    or (f"Job {job}" if job else f"Row {row}")
                ),
                "customer_part_number": part,
                "description": desc,
                "project_type": "Launch",
                "launch_type": launch_type,
                "family": "",
                # Blank rather than a real site would vanish from the plant
                # filter, so unreadable plants are labelled explicitly.
                "plant": _text(cell(PROJECT_COLS["plant"])) or "Unknown",
                "div": _text(cell(PROJECT_COLS["div"])),
                "customer": _text(cell(PROJECT_COLS["customer"])),
                "sales_person": extra.get("sales_person", ""),
                "gate_zero_corp": extra.get("gate_zero_corp", ""),
                "program_manager": pm,
                "job_number": job,
                "qmsi_number": "",
                "qmsi_revision": "",
                "opportunity_number": extra.get("opportunity_number", ""),
                "rpn": int(_num(cell(PROJECT_COLS["rpn"]))),
                "peak_annual_sales": _num(cell(PROJECT_COLS["peak_annual_sales"])),
                "launch_process": process,
                "support_required": extra.get("support_required", ""),
                "launch_risk": _text(cell(PROJECT_COLS["rpn"])),
                "qmsi_capex": _num(cell(PROJECT_COLS["qmsi_capex"])),
                "cer_amount": _num(cell(PROJECT_COLS["cer_amount"])),
                "cer_status": _text(cell(PROJECT_COLS["cer_status"])),
                "cer_number": _text(cell(PROJECT_COLS["cer_number"])),
                "gate_zero_date": gate_zero,
                "ppap_target_date": _date(cell(PROJECT_COLS["ppap_target_date"])),
                "sop_target_date": _date(cell(PROJECT_COLS["sop_target_date"])),
                "project_status": RYG.get(ryg, "Green"),
                "project_phase": _text(cell(PROJECT_COLS["project_phase"])) or "In-Process",
                "prr_count": 0,
                "prr_amount_first_year": _num(cell(PROJECT_COLS["prr_amount_first_year"])),
                "prr_start_date": _date(cell(PROJECT_COLS["prr_start_date"])),
                "prr_end_date": _date(cell(PROJECT_COLS["prr_end_date"])),
                "main_risk_comments": extra.get("main_risk_comments", ""),
                "notes": _text(cell(PROJECT_COLS["notes"])),
            }
        )

        for gate_no, code, name, pc, ac, actc, stc in GATE_MAP:
            # Simple launches carry SL and skip 1-3; full launches the reverse.
            if launch_type == "Simple" and code in ("1", "2", "3"):
                continue
            if launch_type == "Full" and code == "SL":
                continue

            plan = _date(cell(pc)) if pc else None
            adjusted = _date(cell(ac)) if ac else None
            actual = _date(cell(actc)) if actc else None
            status = _text(cell(stc)) if stc else ""

            if plan is None and adjusted is None and actual is None:
                continue

            gates.append(
                {
                    "project_id": pid,
                    "gate_no": gate_no,
                    "gate_code": code,
                    "gate_name": name,
                    "plan_date": plan or adjusted or actual,
                    "adjusted_date": adjusted,
                    "actual_date": actual,
                    "qa_lab_hours": QA_HOURS.get(code, 0.0),
                    "status_override": status,
                }
            )

    if not projects:
        raise ImportError_(
            "No projects found. Check that data starts at row "
            f"{FIRST_DATA_ROW} of '{TRACKER_SHEET}'."
        )

    return pd.DataFrame(projects), pd.DataFrame(gates), warnings


def _iso(v):
    return None if v is None or pd.isna(v) else pd.Timestamp(v).date().isoformat()


def write(projects: pd.DataFrame, gates: pd.DataFrame) -> None:
    """
    Replace the loaded data. This is the most destructive action in the app,
    so it takes the lock and snapshots first - a bad import is then a restore
    away rather than unrecoverable.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    p, g = projects.copy(), gates.copy()
    import db
    import launch_model as lm

    safe_io.backup(reason="pre-import")

    # Drop gates whose project is missing - the foreign key would reject them
    # and take the whole import down with it.
    known = set(p["project_id"])
    g = g[g["project_id"].isin(known)]

    # One transaction: either the whole new dataset lands or none of it.
    db.replace_all(lm.normalise_projects(p), lm.normalise_gates(g))


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    projects, gates, warnings = parse(sys.argv[1])
    write(projects, gates)
    print(f"imported {len(projects)} projects, {len(gates)} gates")
    for w in warnings[:20]:
        print("  warning:", w)
    if len(warnings) > 20:
        print(f"  ... and {len(warnings) - 20} more")


if __name__ == "__main__":
    main()
